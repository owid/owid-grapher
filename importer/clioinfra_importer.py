import sys
import os
import csv
import hashlib
import json
import glob
import unidecode
import lxml.html
import re
import requests
import time
import urllib.parse
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import owid_grapher.wsgi
import django.db
from grapher_admin.models import Entity, DatasetSubcategory, DatasetCategory, Dataset, Source, Variable, VariableType, DataValue
from importer.models import ImportHistory
from country_name_tool.models import CountryName
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from grapher_admin.views import write_dataset_csv
from openpyxl import load_workbook


# we will use the file checksum to check if the downloaded file has changed since we last saw it
def file_checksum(filename, blocksize=2**20):
    m = hashlib.md5()
    with open(filename, "rb") as f:
        while True:
            buffer = f.read(blocksize)
            if not buffer:
                break
            m.update(buffer)
    return m.hexdigest()


def short_unit_extract(unit: str):
    common_short_units = ['$', '£', '€', '%']  # used for extracting short forms of units of measurement
    short_unit = None
    if unit:
        if ' per ' in unit:
            short_form = unit.split(' per ')[0]
            if any(w in short_form for w in common_short_units):
                for x in common_short_units:
                    if x in short_form:
                        short_unit = x
                        break
            else:
                short_unit = short_form
        elif any(x in unit for x in common_short_units):
            for y in common_short_units:
                if y in unit:
                    short_unit = y
                    break
        elif 'percentage' in unit:
            short_unit = '%'
        elif len(unit) < 9:  # this length is sort of arbitrary at this point, taken from the unit 'hectares'
            short_unit = unit
    return short_unit


clioinfra_category_name_in_db = 'Clio-Infra'  # set the name of the root category of all data that will be imported by this script

source_description = {
    'dataPublishedBy': "Clio-Infra",
    'dataPublisherSource': None,
    'retrievedDate': timezone.now().strftime("%d-%B-%y"),
    'additionalInfo': None
}

base_dataverse_url = "https://datasets.socialhistory.org"
# downloaded excel files will be saved here
files_save_location = os.path.join(settings.BASE_DIR, 'data/clioinfra/download_files')

# you should manually put the metadata.csv file in this directory
# the metadata file will only contain the dataset name and its category
# the dataset names should be under Dataset column
# the category names should be under Category column
# the file format is csv
# dataset and category information can be found on clio-infra's dataverse, in the file called topics
metadata_location = os.path.join(settings.BASE_DIR, 'data/clioinfra/metadata')

if not os.path.isfile(os.path.join(metadata_location, 'metadata.csv')):
    print('The metadata.csv file does not exist in %s. Exiting...' % metadata_location)
    sys.exit()

# getting the initial page
page_r = requests.get("https://datasets.socialhistory.org/dataverse/clioinfra?types=datasets")

html = lxml.html.fromstring(page_r.content)
# now getting the number of pages to iterate through
page_lists = html.xpath('//ul[@class="pagination"]/li')
number_of_pages = urllib.parse.urlparse(page_lists[len(page_lists)-1].xpath('.//a/@href')[0])
number_of_pages = int(urllib.parse.parse_qs(number_of_pages.query).get('page', ['1'])[0])

filename_to_pagelink = {}

for page in range(1, number_of_pages + 1):
    page_r = requests.get("https://datasets.socialhistory.org/dataverse/clioinfra?types=datasets&page=%s" % page)
    dataset_page_links = lxml.html.fromstring(page_r.content).xpath('//div[contains(@class, "datasetResult")]/div[@class="card-title-icon-block"]/a/@href')
    for each_link in dataset_page_links:
        # open one dataset's page
        dataset_page = requests.get(base_dataverse_url + each_link)
        # session cookie and view state are needed for downloading the files
        session_cookie = dataset_page.cookies['JSESSIONID']
        page_view_state = lxml.html.fromstring(dataset_page.content).xpath('//input[@name="javax.faces.ViewState"]/@value')[0]

        file_ids = lxml.html.fromstring(dataset_page.content).xpath('//div[@class="ui-datatable-scrollable-body"]//div[@class="button-block"]/a/@id')
        for each_id in file_ids:
            request_headers = {
                'X-Requested-With': 'XMLHttpRequest',
                'Faces-Request': 'partial/ajax',
                'Host': 'datasets.socialhistory.org',
                'Cookie': 'JSESSIONID=%s' % session_cookie
            }

            request_parameters = {
                'datasetForm': 'datasetForm',
                each_id: each_id,
                'datasetForm:options': 1,
                'javax.faces.source': each_id,
                'javax.faces.partial.execute': '@all',
                'javax.faces.ViewState': page_view_state,
                'javax.faces.partial.ajax': 'true'
            }

            file_url = requests.post(base_dataverse_url + each_link, data=request_parameters, headers=request_headers)
            # now downloading the files
            for one_url in file_url:
                download_url = base_dataverse_url + lxml.html.fromstring(one_url.decode('utf-8')).xpath('//redirect/@url')[0]
                file_download = requests.get(download_url, stream=True)
                # we wouldn't want to make too many requests in a short amount of time
                time.sleep(2)
                fname = re.findall("filename=(.+)", file_download.headers['content-disposition'])[0].replace('"', '')
                if 'xlsx' in fname:
                    with open(os.path.join(files_save_location, fname), 'wb') as ff:
                        for chunk in file_download.iter_content(chunk_size=1024):
                            if chunk:
                                ff.write(chunk)
                    filename_to_pagelink[fname] = base_dataverse_url + each_link

dataset_to_category = {}
with open(os.path.join(metadata_location, 'metadata.csv'), encoding='utf-8') as metadata:
    metareader = csv.DictReader(metadata)
    for row in metareader:
        dataset_to_category[row['Dataset']] = row['Category']

import_history = ImportHistory.objects.filter(import_type='clioinfra')

with transaction.atomic():
    new_datasets_list = []
    old_datasets_list = []
    existing_categories = DatasetCategory.objects.values('name')
    existing_categories_list = {item['name'] for item in existing_categories}

    if clioinfra_category_name_in_db not in existing_categories_list:
        the_category = DatasetCategory(name=clioinfra_category_name_in_db, fetcher_autocreated=True)
        the_category.save()
    else:
        the_category = DatasetCategory.objects.get(name=clioinfra_category_name_in_db)

    existing_subcategories = DatasetSubcategory.objects.filter(fk_dst_cat_id=the_category.pk).values('name')
    existing_subcategories_list = {item['name'] for item in existing_subcategories}

    existing_entities = Entity.objects.values('name')
    existing_entities_list = {item['name'].lower() for item in existing_entities}

    country_tool_names = CountryName.objects.all()
    country_tool_names_dict = {}
    for each in country_tool_names:
        country_tool_names_dict[each.country_name.lower()] = each.owid_country

    country_name_entity_ref = {}  # this dict will hold the country names and the appropriate entity object (this is used when saving the variables and their values)

    insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table

    for file in glob.glob(files_save_location + "/*.xlsx"):
        column_to_year = {}
        data_values_dict = {}
        one_file = os.path.basename(file)

        file_imported_before = False
        for oneimport in import_history:
            if json.loads(oneimport.import_state)['file_name'] == one_file:
                file_imported_before = True
                imported_before_hash = json.loads(oneimport.import_state)['file_hash']
        if not file_imported_before:
            # all dataset files contain the word historical
            if 'historical' in one_file.lower():
                print('Processing: %s' % one_file)
                wb = load_workbook(file, read_only=True)

                data_ws = wb['Data']

                column_number = 0  # this will be reset to 0 on each new row
                row_number = 0  # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

                for row in data_ws.rows:
                    row_number += 1
                    for cell in row:

                        column_number += 1

                        if row_number == 1 and column_number == 1:
                            varname = cell.value

                        if row_number == 2 and column_number == 1:
                            varunit = cell.value

                        if row_number == 3 and column_number == 1:
                            # inserting a subcategory and dataset
                            if dataset_to_category[varname] not in existing_subcategories_list:
                                the_subcategory = DatasetSubcategory(name=dataset_to_category[varname],
                                                                     fk_dst_cat_id=the_category)
                                the_subcategory.save()
                                newdataset = Dataset(name='Clio-Infra - %s' % the_subcategory.name,
                                                     description='This is a dataset imported by the automated fetcher',
                                                     namespace='clioinfra', fk_dst_cat_id=the_category,
                                                     fk_dst_subcat_id=the_subcategory)
                                newdataset.save()
                                new_datasets_list.append(newdataset)
                                existing_subcategories_list.add(dataset_to_category[varname])
                            else:
                                the_subcategory = DatasetSubcategory.objects.get(
                                    name=dataset_to_category[varname], fk_dst_cat_id=the_category)
                                newdataset = Dataset.objects.get(name='Clio-Infra - %s' % the_subcategory.name,
                                                                 namespace='clioinfra')
                            source_description['link'] = filename_to_pagelink[one_file]
                            newsource = Source(name=varname,
                                               description=json.dumps(source_description),
                                               datasetId=newdataset.pk)
                            newsource.save()

                            newvariable = Variable(name=varname,
                                                   unit=varunit if
                                                   varunit else '', short_unit=short_unit_extract(varunit),
                                                   description='',
                                                   code=filename_to_pagelink[one_file][filename_to_pagelink[one_file].rfind('/') + 1:], timespan='',
                                                   fk_dst_id=newdataset, fk_var_type_id=VariableType.objects.get(pk=4),
                                                   sourceId=newsource)

                            newvariable.save()

                        if row_number == 3 and column_number > 6:
                            try:
                                column_to_year[column_number] = int(cell.value)
                            except ValueError:
                                pass

                        if row_number > 3:
                            if column_number == 4 and cell.value is not None:
                                countryname = cell.value
                                if countryname not in country_name_entity_ref:
                                    if countryname.lower() in existing_entities_list:
                                        newentity = Entity.objects.get(name=countryname)
                                    elif country_tool_names_dict.get(unidecode.unidecode(countryname.lower()), 0):
                                        newentity = Entity.objects.get(
                                            name=country_tool_names_dict[
                                                unidecode.unidecode(countryname.lower())].owid_name)
                                    else:
                                        newentity = Entity(name=countryname, validated=False)
                                        newentity.save()
                                    country_name_entity_ref[countryname] = newentity

                            if column_number > 6 and cell.value is not None:
                                try:
                                    value = float(cell.value)
                                except ValueError:
                                    continue
                                if data_values_dict.get(country_name_entity_ref[countryname].pk):
                                    data_values_dict[country_name_entity_ref[countryname].pk][column_to_year[column_number]] = str(value)
                                else:
                                    data_values_dict[country_name_entity_ref[countryname].pk] = {}
                                    data_values_dict[country_name_entity_ref[countryname].pk][column_to_year[column_number]] = str(value)

                    column_number = 0

                data_values_tuple_list = []
                for country, data_value in data_values_dict.items():
                    for year, value in data_value.items():
                        data_values_tuple_list.append((value, year, country, newvariable.pk))

                with connection.cursor() as c:
                    c.executemany(insert_string, data_values_tuple_list)

                newimport = ImportHistory(import_type='clioinfra',
                                          import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                          import_notes='Importing file %s' % one_file,
                                          import_state=json.dumps(
                                              {'file_hash': file_checksum(file),
                                               'file_name': one_file
                                               }))
                newimport.save()

        else:
            if imported_before_hash == file_checksum(file):
                print('No updates available for file %s.' % one_file)
            else:
                if 'historical' in one_file.lower():
                    print('Processing: %s' % one_file)
                    wb = load_workbook(file, read_only=True)

                    data_ws = wb['Data']

                    column_number = 0  # this will be reset to 0 on each new row
                    row_number = 0  # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

                    for row in data_ws.rows:
                        row_number += 1
                        for cell in row:

                            column_number += 1

                            if row_number == 1 and column_number == 1:
                                varname = cell.value

                            if row_number == 2 and column_number == 1:
                                varunit = cell.value

                            if row_number == 3 and column_number == 1:
                                # inserting a subcategory and dataset
                                if dataset_to_category[varname] not in existing_subcategories_list:
                                    the_subcategory = DatasetSubcategory(name=dataset_to_category[varname],
                                                                         fk_dst_cat_id=the_category)
                                    the_subcategory.save()
                                    newdataset = Dataset(name='Clio-Infra - %s' % the_subcategory.name,
                                                         description='This is a dataset imported by the automated fetcher',
                                                         namespace='clioinfra', fk_dst_cat_id=the_category,
                                                         fk_dst_subcat_id=the_subcategory)
                                    newdataset.save()
                                    new_datasets_list.append(newdataset)
                                    existing_subcategories_list.add(dataset_to_category[varname])
                                else:
                                    the_subcategory = DatasetSubcategory.objects.get(
                                        name=dataset_to_category[varname], fk_dst_cat_id=the_category)
                                    try:
                                        newdataset = Dataset.objects.get(name='Clio-Infra - %s' % the_subcategory.name,
                                                                     namespace='clioinfra')
                                    except Dataset.DoesNotExist:
                                        newdataset = Dataset.objects.get(name__startswith='Clio-Infra - %s' % the_subcategory.name,
                                                                         namespace='clioinfra')
                                    if newdataset not in old_datasets_list:
                                        old_datasets_list.append(newdataset)

                                newsource = Variable.objects.get(code=filename_to_pagelink[one_file][filename_to_pagelink[one_file].rfind('/') + 1:], fk_dst_id__namespace='clioinfra').sourceId
                                source_description['link'] = filename_to_pagelink[one_file]
                                newsource.description = json.dumps(source_description)
                                newsource.datasetId = newdataset.pk
                                newsource.save()

                                newvariable = Variable.objects.get(code=filename_to_pagelink[one_file][filename_to_pagelink[one_file].rfind('/') + 1:], fk_dst_id__namespace='clioinfra')
                                newvariable.name = varname
                                newvariable.unit = varunit if varunit else ''
                                newvariable.short_unit = short_unit_extract(varunit)
                                newvariable.description = ''
                                newvariable.timespan = ''
                                newvariable.fk_dst_id = newdataset
                                newvariable.fk_var_type_id = VariableType.objects.get(pk=4)
                                newvariable.sourceId = newsource

                                newvariable.save()

                                # Deleting old data values
                                while DataValue.objects.filter(fk_var_id__pk=newvariable.pk).first():
                                    with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                                        c.execute('DELETE FROM %s WHERE fk_var_id = %s LIMIT 10000;' %
                                                  (DataValue._meta.db_table, newvariable.pk))

                            if row_number == 3 and column_number > 6:
                                try:
                                    column_to_year[column_number] = int(cell.value)
                                except ValueError:
                                    pass

                            if row_number > 3:
                                if column_number == 4 and cell.value is not None:
                                    countryname = cell.value
                                    if countryname not in country_name_entity_ref:
                                        if countryname.lower() in existing_entities_list:
                                            newentity = Entity.objects.get(name=countryname)
                                        elif country_tool_names_dict.get(unidecode.unidecode(countryname.lower()), 0):
                                            newentity = Entity.objects.get(
                                                name=country_tool_names_dict[
                                                    unidecode.unidecode(countryname.lower())].owid_name)
                                        else:
                                            newentity = Entity(name=countryname, validated=False)
                                            newentity.save()
                                        country_name_entity_ref[countryname] = newentity

                                if column_number > 6 and cell.value is not None:
                                    try:
                                        value = float(cell.value)
                                    except ValueError:
                                        continue
                                    if data_values_dict.get(country_name_entity_ref[countryname].pk):
                                        data_values_dict[country_name_entity_ref[countryname].pk][
                                            column_to_year[column_number]] = str(value)
                                    else:
                                        data_values_dict[country_name_entity_ref[countryname].pk] = {}
                                        data_values_dict[country_name_entity_ref[countryname].pk][
                                            column_to_year[column_number]] = str(value)

                        column_number = 0

                    data_values_tuple_list = []
                    for country, data_value in data_values_dict.items():
                        for year, value in data_value.items():
                            data_values_tuple_list.append((value, year, country, newvariable.pk))

                    with connection.cursor() as c:
                        c.executemany(insert_string, data_values_tuple_list)

                    newimport = ImportHistory(import_type='clioinfra',
                                              import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                              import_notes='Importing file %s' % one_file,
                                              import_state=json.dumps(
                                                  {'file_hash': file_checksum(file),
                                                   'file_name': one_file
                                                   }))
                    newimport.save()

    for eachdataset in new_datasets_list:
        write_dataset_csv(eachdataset.pk, eachdataset.name, None, 'clioinfra_fetcher', '')
    for eachdataset in old_datasets_list:
        write_dataset_csv(eachdataset.pk, eachdataset.name, eachdataset.name, 'clioinfra_fetcher', '')
