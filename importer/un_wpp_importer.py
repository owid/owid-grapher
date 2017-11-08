import os
import sys
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import owid_grapher.wsgi
import hashlib
import json
import requests
import shutil
import unidecode
from urllib.parse import quote
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from openpyxl import load_workbook
from importer.models import ImportHistory
from country_name_tool.models import CountryName
from grapher_admin.models import Entity, DatasetSubcategory, DatasetCategory, Dataset, Source, Variable, VariableType, DataValue, ChartDimension
from grapher_admin.views import write_dataset_csv


# IMPORTANT: Unlike World Bank and QoG Institute, which have their datasets as a single file,
# the UN makes their data available in many excel files
# These files have different structures: there are currently 6 different layouts for these files
# With the aim of making it easier to parse these files, this script is configured to process one file at a time
# You will need to enter the file name in file_to_parse variable and run the script for each file to be parsed
# Currently, only 2 file structures containing values per each year are supported
# Support for other 4 structures will be implemented when we decide how to deal with time intervals and not only years


# Enter the name of the file to parse (include the xls or xlsx extension)
file_to_parse = 'WPP2017_INT_F02C_1_ANNUAL_POPULATION_INDICATORS_DEPENDENCY_RATIOS_BOTH_SEXES.xlsx'

un_wpp_data_page_url = 'https://esa.un.org/unpd/wpp/Download/Standard/Population/'
un_wpp_root_url = 'https://esa.un.org/'
wpp_downloads_save_location = settings.BASE_DIR + '/data/un_wpp_downloads/'

# create a directory for holding the downloads
# if the directory exists, delete it and recreate it

source_description = {
    'dataPublishedBy': "United Nations, Department of Economic and Social Affairs, Population Division (2017). World Population Prospects: The 2017 Revision, DVD Edition.",
    'dataPublisherSource': None,
    'link': "https://esa.un.org/unpd/wpp/Download/Standard/Population/",
    'retrievedDate': timezone.now().strftime("%d-%B-%y")
}

# we will use the file checksum to check if the downloaded file has changed since we last saw it
def file_checksum(filename, blocksize=2**20):
    m = hashlib.md5()
    with open(filename, "rb") as f:
        while True:
            buffer = f.read(blocksize)
            if not buffer:
                break
            m.update(buffer)
    return m.hexdigest()

if not os.path.exists(wpp_downloads_save_location):
    os.makedirs(wpp_downloads_save_location)
else:
    shutil.rmtree(wpp_downloads_save_location)
    os.makedirs(wpp_downloads_save_location)


def short_unit_extract(unit: str):
    common_short_units = ['$', '£', '€', '%']  # used for extracting short forms of units of measurement
    short_unit = None
    if unit:
        if ' per ' in unit:
            short_form = unit.split(' per ')[0]
            if any(w in short_form for w in common_short_units):
                for x in common_short_units:
                    if x in short_form:
                        short_unit = x
                        break
            else:
                short_unit = short_form
        elif any(x in unit for x in common_short_units):
            for y in common_short_units:
                if y in unit:
                    short_unit = y
                    break
        elif len(unit) < 9:  # this length is sort of arbitrary at this point, taken from the unit 'hectares'
            short_unit = unit
    return short_unit


def process_entities(country_names_dictionary):
    existing_entities = Entity.objects.values('name')
    existing_entities_list = {item['name'].lower() for item in existing_entities}

    country_tool_names = CountryName.objects.all()
    country_tool_names_dict = {}

    for each_country in country_tool_names:
        country_tool_names_dict[each_country.country_name.lower()] = each_country.owid_country

    c_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

    for c_code, country_name in country_names_dictionary.items():
        if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
            newentity = Entity.objects.get(
                name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
        elif country_name.lower() in existing_entities_list:
            newentity = Entity.objects.get(name__iexact=country_name)
        else:
            newentity = Entity(name=country_name, validated=False)
            newentity.save()
        c_name_entity_ref[c_code] = newentity

    return c_name_entity_ref


request_header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
r = requests.get(un_wpp_data_page_url, headers=request_header)
dataset_info = {}  # will contain filename, and category name and description for the file
file_found = False

if r.ok:
    if 'var filesArray' not in r.text:
        sys.exit('Could not find the array of files. Exiting now...')
    else:
        # getting the json array containing all the xlsx files
        files_array = r.text[r.text.index('[', r.text.index('var filesArray')):r.text.index(']', r.text.index('var filesArray')) + 1]
        files_dict = json.loads(files_array)
        for each in files_dict:
            # we want to get only the excel files in the following groups
            # Population, Mortality, Fertility, Migration and Interpolated indicators
            # we also want to exclude the abridged files
            if '.xlsx' in each['File1_Path'] and '1_Indicators (Standard)' in each['File1_Path'] and 'ABRIDGED' not in each['File1_Path'] and file_to_parse in each['File1_Path']:
                r = requests.get(un_wpp_root_url + quote(each['File1_Path']), stream=True, headers=request_header)
                file_found = True
                if r.ok:
                    with open(wpp_downloads_save_location + os.path.basename(un_wpp_root_url + each['File1_Path']), 'wb') as out_file:
                        shutil.copyfileobj(r.raw, out_file)
                    dataset_info['filename'] = file_to_parse
                    dataset_info['category'] = '%s - %s' % (each['MajorGroup'], each['SubGroup'])
                    dataset_info['description'] = each['Description']
                else:
                    sys.exit('Could not download %s. Exiting now...' % each['File1_Path'])
else:
    sys.exit('Could not download the contents of the json file from the server. Exiting now...')

if not file_found:
    sys.exit('The file you requested was not found on the server. Exiting now...')

un_wpp_category_name_in_db = 'United Nations World Population Prospects'  # set the name of the root category of all data that will be imported by this script


# now we determine the structure type the file before starting to parse it
# the for loop below will update the files_catalog dictionary
# there are 6 structure types: see the below for loop, and only 2 of them supported at the moment
# the for loop below will also insert new entities and will construct a dict of entity names for using later

with transaction.atomic():

    wb = load_workbook(os.path.join(wpp_downloads_save_location, file_to_parse), read_only=True)

    ws_names = wb.get_sheet_names()
    column_number = 0
    row_number = 0
    horizontal = []  # used to determine how the years are displayed in the file (e.g.: in columns or rows)
    vertical = []
    found_country_col = False
    country_names_dict = {}

    for row in wb[ws_names[0]]:
        row_number += 1
        for cell in row:
            column_number += 1
            if row_number == 17:
                if found_country_col:
                    horizontal.append(cell.value)
                if 'Country code' in cell.value:
                    year_col_number = column_number + 1
                    found_country_col = True
            if row_number > 17 and row_number < 24:
                if column_number == year_col_number:
                    vertical.append(cell.value)

            if row_number > 17:
                if column_number == 3:
                    if cell.value:
                        the_country_name = cell.value
                if column_number == 5:
                    if 'Serbia' == the_country_name:
                        the_country_name = 'Serbia (including Kosovo)'
                    if 'Guadeloupe' == the_country_name:
                        the_country_name = 'Guadeloupe (including Saint-Barthélemy and Saint-Martin)'
                    if 'United Republic of Tanzania' == the_country_name:
                        the_country_name = 'Tanzania'
                    if 'China' == the_country_name:
                        the_country_name = 'China'
                    if 'China, Hong Kong SAR' == the_country_name:
                        the_country_name = 'Hong Kong'
                    if 'China, Macao SAR' == the_country_name:
                        the_country_name = 'Macao'
                    if 'China, Taiwan Province of China' == the_country_name:
                        the_country_name = 'Taiwan'
                    if 'State of Palestine' == the_country_name:
                        the_country_name = 'Palestine'
                    if 'Czechia' == the_country_name:
                        the_country_name = 'Czech Republic'
                    if 'Republic of Moldova' == the_country_name:
                        the_country_name = 'Moldova'
                    if 'TFYR Macedonia' == the_country_name:
                        the_country_name = 'Macedonia'
                    if 'NORTHERN AMERICA' == the_country_name:
                        the_country_name = 'Northern America'
                    if the_country_name == 'Micronesia':
                        the_country_name = 'Micronesia (region)'
                    if the_country_name == 'Micronesia (Fed. States of)':
                        the_country_name = 'Micronesia (country)'

                    if cell.value:
                        if cell.value not in country_names_dict:
                            country_names_dict[cell.value] = the_country_name

        column_number = 0

    if '1950-1955' in str(horizontal[0]) and '1955-1960' in str(horizontal[1]):
        dataset_info['structure'] = 1
    elif "1950" == str(horizontal[0]) and "1955" == str(horizontal[1]):
        dataset_info['structure'] = 2
    elif "1950" == str(horizontal[0]) and "1951" == str(horizontal[1]):
        dataset_info['structure'] = 3

    if '1950-1955' in str(vertical[0]) and '1955-1960' in str(vertical[1]):
        dataset_info['structure'] = 4
    elif "1950" == str(vertical[0]) and "1955" == str(vertical[1]):
        dataset_info['structure'] = 5
    elif "1950" == str(vertical[0]) and "1951" == str(vertical[1]):
        dataset_info['structure'] = 6

    if dataset_info['structure'] != 6 and dataset_info['structure'] != 3:
        sys.exit('This file structure is not supported yet!')

    import_history = ImportHistory.objects.filter(import_type='unwpp')
    file_imported_before = False
    for oneimport in import_history:
        if json.loads(oneimport.import_state)['file_name'] == file_to_parse:
            file_imported_before = True
            imported_before_hash = json.loads(oneimport.import_state)['file_hash']

    # if unwpp imports for this file were never performed
    if not file_imported_before:

        country_name_entity_ref = process_entities(country_names_dict)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if un_wpp_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=un_wpp_category_name_in_db, fetcher_autocreated=True)
            the_category.save()

        else:
            the_category = DatasetCategory.objects.get(name=un_wpp_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(fk_dst_cat_id=the_category.pk).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        the_subcategory_name = dataset_info['category']
        if the_subcategory_name not in existing_subcategories_list:
            the_subcategory = DatasetSubcategory(name=the_subcategory_name, fk_dst_cat_id=the_category)
            the_subcategory.save()
        else:
            the_subcategory = DatasetSubcategory.objects.get(name=the_subcategory_name, fk_dst_cat_id=the_category)

        wb = load_workbook(os.path.join(wpp_downloads_save_location, file_to_parse), read_only=True)
        sheets = wb.get_sheet_names()
        sheets.remove('NOTES')  # we don't need this sheet

        if dataset_info['structure'] == 6:
            dataset_saved = False
            for sheet in sheets:
                variables_saved = False
                column_number = 0
                row_number = 0
                var_to_add_dict = {}
                column_var_dict = {}
                data_values_tuple_list = []

                insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                                timespan = variant[variant.index(', ') + 2:]
                        if row_number == 16:
                            if cell.value:
                                main_var_name = cell.value
                        if row_number == 17:
                            if column_number > 6:
                                if cell.value:
                                    var_to_add_dict[column_number] = '%s: %s - %s' % (variant, main_var_name, cell.value)

                        if row_number == 18:
                            if not dataset_saved:
                                newdataset = Dataset(name='UN WPP - %s' % dataset_name,
                                                     description='This is a dataset imported by the automated fetcher',
                                                     namespace='unwpp', fk_dst_cat_id=the_category,
                                                     fk_dst_subcat_id=the_subcategory)
                                newdataset.save()
                                dataset_saved = True
                                source_description['additionalInfo'] = dataset_info['description']
                                newsource = Source(name='United Nations – Population Division (2017 Revision)',
                                                   description=json.dumps(source_description),
                                                   datasetId=newdataset.pk)
                                newsource.save()

                            if not variables_saved:
                                for columnnum, varname in var_to_add_dict.items():
                                    if '(' not in varname:
                                        unit_of_measure = ''
                                    else:
                                        unit_of_measure = varname[varname.index('('):varname.index(')') + 1].replace('(', '').replace(')','')
                                        s_unit = short_unit_extract(unit_of_measure)
                                    newvariable = Variable(name=varname,
                                                           unit=unit_of_measure,
                                                           short_unit=s_unit,
                                                           description='',
                                                           code=None,
                                                           timespan=timespan,
                                                           fk_dst_id=newdataset,
                                                           fk_var_type_id=VariableType.objects.get(pk=4),
                                                           sourceId=newsource)
                                    newvariable.save()

                                    column_var_dict[columnnum] = newvariable

                                variables_saved = True

                        if row_number > 17:
                            if column_number == 5:
                                country_code = cell.value
                            if column_number == 6:
                                year = cell.value
                            if column_number > 6:
                                if cell.value:
                                    try:
                                        the_value = float(cell.value)
                                        data_values_tuple_list.append((cell.value, year,
                                                                        country_name_entity_ref[country_code].pk,
                                                                        column_var_dict[column_number].pk))
                                    except ValueError:
                                        # the cell value does not contain a number
                                        pass

                    column_number = 0

                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                        with connection.cursor() as c:
                            c.executemany(insert_string, data_values_tuple_list)
                        data_values_tuple_list = []

                if len(data_values_tuple_list):  # insert any leftover data_values
                    with connection.cursor() as c:
                        c.executemany(insert_string, data_values_tuple_list)

        elif dataset_info['structure'] == 3:
            dataset_saved = False
            for sheet in sheets:
                variables_saved = False
                column_number = 0
                row_number = 0
                column_to_year = {}
                data_values_tuple_list = []

                insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                                timespan = variant[variant.index(', ') + 2:]
                        if row_number == 16:
                            if cell.value:
                                var_name = cell.value
                        if row_number == 17:
                            if column_number > 5:
                                if cell.value:
                                    column_to_year[column_number] = cell.value

                        if row_number == 18:
                            if not dataset_saved:
                                newdataset = Dataset(name='UN WPP - %s' % dataset_name,
                                                     description='This is a dataset imported by the automated fetcher',
                                                     namespace='unwpp', fk_dst_cat_id=the_category,
                                                     fk_dst_subcat_id=the_subcategory)
                                newdataset.save()
                                dataset_saved = True
                                source_description['additionalInfo'] = dataset_info['description']
                                newsource = Source(name='United Nations – Population Division (2017 Revision)',
                                                   description=json.dumps(source_description),
                                                   datasetId=newdataset.pk)
                                newsource.save()

                            if not variables_saved:
                                if '(' not in var_name:
                                    unit_of_measure = ''
                                else:
                                    unit_of_measure = var_name[var_name.index('('):var_name.index(')') + 1].replace('(', '').replace(')','')
                                    s_unit = short_unit_extract(unit_of_measure)
                                newvariable = Variable(name='%s: %s' % (variant, var_name),
                                                       unit=unit_of_measure,
                                                       short_unit=s_unit,
                                                       description='',
                                                       code=None,
                                                       timespan=timespan,
                                                       fk_dst_id=newdataset,
                                                       fk_var_type_id=VariableType.objects.get(pk=4),
                                                       sourceId=newsource)
                                newvariable.save()

                                variables_saved = True

                        if row_number > 17:
                            if column_number == 5:
                                country_code = cell.value
                            if column_number > 5:
                                if cell.value:
                                    try:
                                        the_value = float(cell.value)
                                        data_values_tuple_list.append((cell.value, column_to_year[column_number],
                                                                       country_name_entity_ref[country_code].pk,
                                                                       newvariable.pk))
                                    except ValueError:
                                        # the cell value does not contain a number
                                        pass

                    column_number = 0

                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                        with connection.cursor() as c:
                            c.executemany(insert_string, data_values_tuple_list)
                        data_values_tuple_list = []

                if len(data_values_tuple_list):  # insert any leftover data_values
                    with connection.cursor() as c:
                        c.executemany(insert_string, data_values_tuple_list)

        newimport = ImportHistory(import_type='unwpp',
                                      import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                      import_notes='Importing file %s' % file_to_parse,
                                      import_state=json.dumps(
                                      {'file_hash': file_checksum(os.path.join(wpp_downloads_save_location, file_to_parse)),
                                       'file_name': file_to_parse
                                       }))
        newimport.save()
        write_dataset_csv(newdataset.pk, newdataset.name, None, 'unwpp_fetcher', '')
    else:

        if imported_before_hash == file_checksum(os.path.join(wpp_downloads_save_location, file_to_parse)):
            sys.exit('No updates available.')

        country_name_entity_ref = process_entities(country_names_dict)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if un_wpp_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=un_wpp_category_name_in_db, fetcher_autocreated=True)
            the_category.save()

        else:
            the_category = DatasetCategory.objects.get(name=un_wpp_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(fk_dst_cat_id=the_category.pk).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        the_subcategory_name = dataset_info['category']
        if the_subcategory_name not in existing_subcategories_list:
            the_subcategory = DatasetSubcategory(name=the_subcategory_name, fk_dst_cat_id=the_category)
            the_subcategory.save()
        else:
            the_subcategory = DatasetSubcategory.objects.get(name=the_subcategory_name, fk_dst_cat_id=the_category)

        wb = load_workbook(os.path.join(wpp_downloads_save_location, file_to_parse), read_only=True)
        sheets = wb.get_sheet_names()
        sheets.remove('NOTES')  # we don't need this sheet

        if dataset_info['structure'] == 6:
            new_variables = []
            for sheet in sheets:
                column_number = 0
                row_number = 0

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                        if row_number == 16:
                            if cell.value:
                                main_var_name = cell.value
                        if row_number == 17:
                            if column_number > 6:
                                if cell.value:
                                    new_variables.append('%s: %s - %s' % (variant, main_var_name, cell.value))
                        if row_number > 17:
                            break
                    column_number = 0

            thedataset = Dataset.objects.get(name='UN WPP - %s' % dataset_name, namespace__contains='unwpp')
            available_variables = Variable.objects.filter(
                fk_dst_id=thedataset)
            available_variables_list = []

            for each in available_variables.values('name'):
                available_variables_list.append(each['name'])

            chart_dimension_vars = ChartDimension.objects.all().values('variableId').distinct()
            chart_dimension_vars_list = {item['variableId'] for item in chart_dimension_vars}
            existing_variables_ids = [item['id'] for item in available_variables.values('id')]
            existing_variables_id_name = {item['id']: item['name'] for item in
                                          available_variables.values('id', 'name')}
            existing_variables_name_id = {item['name']: item['id'] for item in
                                          available_variables.values('id', 'name')}

            vars_being_used = []  # we will not be deleting any variables that are currently being used by charts
            for each_var in existing_variables_ids:
                if each_var in chart_dimension_vars_list:
                    vars_being_used.append(existing_variables_id_name[each_var])

            vars_to_add = list(set(new_variables).difference(available_variables_list))
            vars_to_delete = list(set(available_variables_list).difference(new_variables))

            for each in vars_to_delete:
                if each not in vars_being_used:
                    while DataValue.objects.filter(fk_var_id__pk=existing_variables_name_id[each]).first():
                        with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                            c.execute('DELETE FROM %s WHERE fk_var_id = %s LIMIT 10000;' %
                                      (DataValue._meta.db_table, existing_variables_name_id[each]))
                    Variable.objects.get(name=each, fk_dst_id=thedataset).delete()

            dataset_saved = False

            for sheet in sheets:
                variables_saved = False
                column_number = 0
                row_number = 0
                var_to_add_dict = {}
                column_var_dict = {}
                data_values_tuple_list = []

                insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                                timespan = variant[variant.index(', ') + 2:]
                        if row_number == 16:
                            if cell.value:
                                main_var_name = cell.value
                        if row_number == 17:
                            if column_number > 6:
                                if cell.value:
                                    var_to_add_dict[column_number] = '%s: %s - %s' % (variant, main_var_name, cell.value)

                        if row_number == 18:
                            if not dataset_saved:
                                newdataset = Dataset.objects.get(name='UN WPP - %s' % dataset_name, namespace__contains='unwpp')
                                newdataset.fk_dst_cat_id = the_category
                                newdataset.fk_dst_subcat_id = the_subcategory
                                newdataset.save()
                                dataset_saved = True

                                source_description['additionalInfo'] = dataset_info['description']
                                newsource = Source.objects.get(datasetId=newdataset.pk)
                                newsource.description = json.dumps(source_description)
                                newsource.save()

                            if not variables_saved:

                                for columnnum, varname in var_to_add_dict.items():
                                    if varname in vars_to_add:
                                        if '(' not in varname:
                                            unit_of_measure = ''
                                        else:
                                            unit_of_measure = varname[varname.index('('):varname.index(')') + 1].replace('(', '').replace(')','')
                                            s_unit = short_unit_extract(unit_of_measure)
                                        newvariable = Variable(name=varname,
                                                               unit=unit_of_measure,
                                                               short_unit=s_unit,
                                                               description='',
                                                               code=None,
                                                               timespan=timespan,
                                                               fk_dst_id=newdataset,
                                                               fk_var_type_id=VariableType.objects.get(pk=4),
                                                               sourceId=newsource)
                                        newvariable.save()
                                    else:
                                        newvariable = Variable.objects.get(name=varname, fk_dst_id=newdataset)
                                        while DataValue.objects.filter(
                                            fk_var_id__pk=existing_variables_name_id[varname]).first():
                                            with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                                                c.execute('DELETE FROM %s WHERE fk_var_id = %s LIMIT 10000;' %
                                                          (DataValue._meta.db_table, existing_variables_name_id[varname]))

                                    column_var_dict[columnnum] = newvariable

                                variables_saved = True

                        if row_number > 17:
                            if column_number == 5:
                                country_code = cell.value
                            if column_number == 6:
                                year = cell.value
                            if column_number > 6:
                                if cell.value:
                                    try:
                                        the_value = float(cell.value)
                                        data_values_tuple_list.append((cell.value, year,
                                                                        country_name_entity_ref[country_code].pk,
                                                                        column_var_dict[column_number].pk))
                                    except ValueError:
                                        # the cell value does not contain a number
                                        pass

                    column_number = 0

                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                        with connection.cursor() as c:
                            c.executemany(insert_string, data_values_tuple_list)
                        data_values_tuple_list = []

                if len(data_values_tuple_list):  # insert any leftover data_values
                    with connection.cursor() as c:
                        c.executemany(insert_string, data_values_tuple_list)

        elif dataset_info['structure'] == 3:
            new_variables = []
            for sheet in sheets:
                column_number = 0
                row_number = 0

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                        if row_number == 16:
                            if cell.value:
                                var_name = cell.value
                                new_variables.append('%s: %s' % (variant, var_name))
                        if row_number > 16:
                            break
                    column_number = 0

            thedataset = Dataset.objects.get(name='UN WPP - %s' % dataset_name, namespace__contains='unwpp')
            available_variables = Variable.objects.filter(
                fk_dst_id=thedataset)
            available_variables_list = []

            for each in available_variables.values('name'):
                available_variables_list.append(each['name'])

            chart_dimension_vars = ChartDimension.objects.all().values('variableId').distinct()
            chart_dimension_vars_list = {item['variableId'] for item in chart_dimension_vars}
            existing_variables_ids = [item['id'] for item in available_variables.values('id')]
            existing_variables_id_name = {item['id']: item['name'] for item in
                                          available_variables.values('id', 'name')}
            existing_variables_name_id = {item['name']: item['id'] for item in
                                          available_variables.values('id', 'name')}

            vars_being_used = []  # we will not be deleting any variables that are currently being used by charts
            for each_var in existing_variables_ids:
                if each_var in chart_dimension_vars_list:
                    vars_being_used.append(existing_variables_id_name[each_var])

            vars_to_add = list(set(new_variables).difference(available_variables_list))
            vars_to_delete = list(set(available_variables_list).difference(new_variables))

            for each in vars_to_delete:
                if each not in vars_being_used:
                    while DataValue.objects.filter(fk_var_id__pk=existing_variables_name_id[each]).first():
                        with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                            c.execute('DELETE FROM %s WHERE fk_var_id = %s LIMIT 10000;' %
                                      (DataValue._meta.db_table, existing_variables_name_id[each]))
                    Variable.objects.get(name=each, fk_dst_id=thedataset).delete()

            dataset_saved = False

            for sheet in sheets:
                variables_saved = False
                column_number = 0
                row_number = 0
                column_to_year = {}
                data_values_tuple_list = []

                insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table

                for row in wb[sheet]:
                    row_number += 1
                    for cell in row:
                        column_number += 1

                        if row_number == 10:
                            if cell.value:
                                dataset_name = cell.value[cell.value.index(': ') + 2:]  # taking the dataset name
                        if row_number == 11:
                            if cell.value:
                                variant = cell.value
                                timespan = variant[variant.index(', ') + 2:]
                        if row_number == 16:
                            if cell.value:
                                var_name = '%s: %s' % (variant, cell.value)
                        if row_number == 17:
                            if column_number > 5:
                                if cell.value:
                                    column_to_year[column_number] = cell.value

                        if row_number == 18:
                            if not dataset_saved:
                                newdataset = Dataset.objects.get(name='UN WPP - %s' % dataset_name, namespace__contains='unwpp')
                                newdataset.fk_dst_cat_id = the_category
                                newdataset.fk_dst_subcat_id = the_subcategory
                                newdataset.save()
                                dataset_saved = True

                                source_description['additionalInfo'] = dataset_info['description']
                                newsource = Source.objects.get(datasetId=newdataset.pk)
                                newsource.description = json.dumps(source_description)
                                newsource.save()

                            if not variables_saved:
                                if var_name in vars_to_add:
                                    if '(' not in var_name:
                                        unit_of_measure = ''
                                    else:
                                        unit_of_measure = var_name[var_name.index('('):var_name.index(')') + 1].replace('(', '').replace(')','')
                                        s_unit = short_unit_extract(unit_of_measure)
                                    newvariable = Variable(name=var_name,
                                                           unit=unit_of_measure,
                                                           short_unit=s_unit,
                                                           description='',
                                                           code=None,
                                                           timespan=timespan,
                                                           fk_dst_id=newdataset,
                                                           fk_var_type_id=VariableType.objects.get(pk=4),
                                                           sourceId=newsource)
                                    newvariable.save()
                                else:
                                    newvariable = Variable.objects.get(name=var_name, fk_dst_id=newdataset)
                                    while DataValue.objects.filter(
                                        fk_var_id__pk=existing_variables_name_id[var_name]).first():
                                        with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                                            c.execute('DELETE FROM %s WHERE fk_var_id = %s LIMIT 10000;' %
                                                      (DataValue._meta.db_table, existing_variables_name_id[var_name]))

                                variables_saved = True

                        if row_number > 17:
                            if column_number == 5:
                                country_code = cell.value
                            if column_number > 5:
                                if cell.value:
                                    try:
                                        the_value = float(cell.value)
                                        data_values_tuple_list.append((cell.value, column_to_year[column_number],
                                                                        country_name_entity_ref[country_code].pk,
                                                                        newvariable.pk))
                                    except ValueError:
                                        # the cell value does not contain a number
                                        pass

                    column_number = 0

                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                        with connection.cursor() as c:
                            c.executemany(insert_string, data_values_tuple_list)
                        data_values_tuple_list = []

                if len(data_values_tuple_list):  # insert any leftover data_values
                    with connection.cursor() as c:
                        c.executemany(insert_string, data_values_tuple_list)


        newimport = ImportHistory(import_type='unwpp',
                                  import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Updating the dataset from file %s' % file_to_parse,
                                  import_state=json.dumps(
                                      {'file_hash': file_checksum(
                                          os.path.join(wpp_downloads_save_location, file_to_parse)),
                                       'file_name': file_to_parse
                                       }))
        newimport.save()
        write_dataset_csv(newdataset.pk, newdataset.name, newdataset.name, 'unwpp_fetcher', '')
