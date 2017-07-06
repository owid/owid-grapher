import sys
import os
import hashlib
import json
import logging
import requests
import unidecode
import shutil
import time
import zipfile
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import owid_grapher.wsgi
from openpyxl import load_workbook
from grapher_admin.models import Entity, DatasetSubcategory, DatasetCategory, Dataset, Source, Variable, VariableType, DataValue
from importer.models import ImportHistory, AdditionalCountryInfo
from country_name_tool.models import CountryName
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from django.urls import reverse


# we will use the file checksum to check if the downloaded file has changed since we last saw it
def file_checksum(filename, blocksize=2**20):
    m = hashlib.md5()
    with open(filename, "rb") as f:
        while True:
            buffer = f.read(blocksize)
            if not buffer:
                break
            m.update(buffer)
    return m.hexdigest()

source_template = '<table>' \
                    '<tr>' \
                        '<td>Variable category</td>' \
                        '<td>%s</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Unit of measure</td>' \
                        '<td>%s</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Variable code in original source</td>' \
                        '<td>%s</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Data published by</td>' \
                        '<td>%s</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Data publisher\'s source</td>' \
                        '<td>%s</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Link</td>' \
                        '<td>http://data.worldbank.org/data-catalog/world-development-indicators</td>' \
                    '</tr>' \
                    '<tr>' \
                        '<td>Retrieved</td>' \
                        '<td>' + timezone.now().strftime("%d-%B-%y") +'</td>' \
                    '</tr>' \
                  '</table>' \
                  '<div class="datasource-additional">' \
                    '<p>' \
                    '<b><i style="text-decoration: underline;">Additional information as provided by the source</i></b><br>' \
                    '<i style="text-decoration: underline;">Definitions and characteristics of countries and other territories:</i> <a href="' + reverse("servewdicountryinfo") + '">WDI_Country_info.xlsx</a><br>' \
                    '%s' \
                    '%s' \
                    '%s' \
                    '%s' \
                    '%s' \
                    '%s' \
                    '</p>' \
                  '</div>'


wdi_zip_file_url = 'http://databank.worldbank.org/data/download/WDI_excel.zip'
wdi_downloads_save_location = settings.BASE_DIR + '/data/wdi_downloads/'

# create a directory for holding the downloads
# if the directory exists, delete it and recreate it

if not os.path.exists(wdi_downloads_save_location):
    os.makedirs(wdi_downloads_save_location)
#else:
#    shutil.rmtree(wdi_downloads_save_location)
#    os.makedirs(wdi_downloads_save_location)

logger = logging.getLogger('importer')
start_time = time.time()

logger.info("Getting the zip file")
request_header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
r = requests.get(wdi_zip_file_url, stream=True, headers=request_header)
if r.ok:
    with open(wdi_downloads_save_location + 'wdi.zip', 'wb') as out_file:
        shutil.copyfileobj(r.raw, out_file)
    logger.info("Saved the zip file to disk.")
    z = zipfile.ZipFile(wdi_downloads_save_location + 'wdi.zip')
    excel_filename = wdi_downloads_save_location + z.namelist()[0]  # there should be only one file inside the zipfile, so we will load that one
    z.extractall(wdi_downloads_save_location)
    r = None  # we do not need the request anymore
    logger.info("Successfully extracted the zip file")
else:
    logger.error("The file could not be downloaded. Stopping the script...")
    sys.exit("Could not download file.")

wdi_category_name_in_db = 'World Development Indicators'  # set the name of the root category of all data that will be imported by this script

import_history = ImportHistory.objects.filter(import_type='wdi')

#excel_filename = wdi_downloads_save_location + "WDIEXCEL.xlsx"

with transaction.atomic():
    # if wdi imports were never performed
    if not import_history:
        logger.info("This is the very first WDI data import.")

        wb = load_workbook(excel_filename, read_only=True)

        series_ws = wb['Series']
        data_ws = wb['Data']
        country_ws = wb['Country']

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0   # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        global_cat = {}  # global catalog of indicators

        # data in the worksheets is not loaded into memory at once, that causes RAM to quickly fill up
        # instead, we go through each row and cell one-by-one, looking at each piece of data separately
        # this has the disadvantage of needing to traverse the worksheet several times, if we need to look up some rows/cells again

        for row in series_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        global_cat[cell.value] = {}
                        indicatordict = global_cat[cell.value]
                    if column_number == 2:
                        indicatordict['category'] = cell.value.split(':')[0]
                    if column_number == 3:
                        indicatordict['name'] = cell.value
                    if column_number == 5:
                        indicatordict['description'] = cell.value
                    if column_number == 6:
                        if cell.value:
                            indicatordict['unitofmeasure'] = cell.value
                        else:
                            if '(' not in indicatordict['name']:
                                indicatordict['unitofmeasure'] = ''
                            else:
                                indicatordict['unitofmeasure'] = indicatordict['name'][indicatordict['name'].index('('):-1].replace('(', '').replace(')','')
                    if column_number == 11:
                        if cell.value:
                            indicatordict['limitations'] = cell.value
                        else:
                            indicatordict['limitations'] = ''
                    if column_number == 12:
                        if cell.value:
                            indicatordict['sourcenotes'] = cell.value
                        else:
                            indicatordict['sourcenotes'] = ''
                    if column_number == 13:
                        if cell.value:
                            indicatordict['comments'] = cell.value
                        else:
                            indicatordict['comments'] = ''
                    if column_number == 14:
                        indicatordict['source'] = cell.value
                    if column_number == 15:
                        if cell.value:
                            indicatordict['concept'] = cell.value
                        else:
                            indicatordict['concept'] = ''
                    if column_number == 17:
                        if cell.value:
                            indicatordict['sourcelinks'] = cell.value
                        else:
                            indicatordict['sourcelinks'] = ''
                    if column_number == 18:
                        if cell.value:
                            indicatordict['weblinks'] = cell.value
                        else:
                            indicatordict['weblinks'] = ''
                    indicatordict['saved'] = False

            column_number = 0

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if wdi_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=wdi_category_name_in_db)
            the_category.save()
            logger.info("Inserting a category %s." % wdi_category_name_in_db.encode('utf8'))

        else:
            the_category = DatasetCategory.objects.get(name=wdi_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(fk_dst_cat_id=the_category.pk).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        wdi_categories_list = []

        for key, value in category_vars.items():
            wdi_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = DatasetSubcategory(name=key, fk_dst_cat_id=the_category)
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'] for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        row_number = 0
        for row in country_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        country_code = cell.value
                    if column_number == 3:
                        country_name = cell.value
                    if column_number == 7:
                        country_special_notes = cell.value
                    if column_number == 8:
                        country_region = cell.value
                    if column_number == 9:
                        country_income_group = cell.value
                    if column_number == 24:
                        country_latest_census = cell.value
                    if column_number == 25:
                        country_latest_survey = cell.value
                    if column_number == 26:
                        country_recent_income_source = cell.value
                    if column_number == 31:
                        entity_info = AdditionalCountryInfo()
                        entity_info.country_code = country_code
                        entity_info.country_name = country_name
                        entity_info.country_wb_region = country_region
                        entity_info.country_wb_income_group = country_income_group
                        entity_info.country_special_notes = country_special_notes
                        entity_info.country_latest_census = country_latest_census
                        entity_info.country_latest_survey = country_latest_survey
                        entity_info.country_recent_income_source = country_recent_income_source
                        entity_info.save()
                        if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                            newentity = Entity.objects.get(name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                        elif country_name in existing_entities_list:
                            newentity = Entity.objects.get(name=country_name)
                        else:
                            newentity = Entity(name=country_name, validated=False)
                            newentity.save()
                            logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                        country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []

        for category in wdi_categories_list:
            newdataset = Dataset(name='World Development Indicators - ' + category,
                                 description='This is a dataset imported by the automated fetcher',
                                 namespace='wdi', fk_dst_cat_id=the_category,
                                 fk_dst_subcat_id=DatasetSubcategory.objects.get(name=category, fk_dst_cat_id=the_category))
            newdataset.save()
            logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            row_number = 0
            for row in data_ws.rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 1:
                            country_name = cell.value
                        if column_number == 2:
                            country_code = cell.value
                        if column_number == 3:
                            indicator_name = cell.value
                        if column_number == 4:
                            indicator_code = cell.value
                        if column_number > 4 and column_number < last_available_year - 1960 + 5:
                            if cell.value:
                                data_values.append({'value': cell.value, 'year': 1960 - 5 + column_number})
                        if column_number > 4 and column_number == last_available_year - 1960 + 5:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    if not global_cat[indicator_code]['saved']:
                                        newsource = Source(name='World Bank – WDI: ' + global_cat[indicator_code]['name'],
                                                           description=source_template %
                                                           (global_cat[indicator_code]['category'],
                                                            global_cat[indicator_code]['unitofmeasure'],
                                                            indicator_code,
                                                            'World Bank – World Development Indicators',
                                                            global_cat[indicator_code]['source'],
                                                            '<i style="text-decoration: underline;">Limitations and exceptions: </i><br>' +
                                                            global_cat[indicator_code]['limitations'] + '<br>' if global_cat[indicator_code]['limitations'] else '',
                                                            '<i style="text-decoration: underline;">Notes from original source: </i><br>' +
                                                            global_cat[indicator_code]['sourcenotes'] + '<br>' if global_cat[indicator_code]['sourcenotes'] else '',
                                                            '<i style="text-decoration: underline;">General comments: </i><br>' + global_cat[indicator_code][
                                                                'comments'] + '<br>' if global_cat[indicator_code]['comments'] else '',
                                                            '<i style="text-decoration: underline;">Statistical concept and methodology: </i><br>' +
                                                            global_cat[indicator_code]['concept'] + '<br>' if global_cat[indicator_code]['concept'] else '',
                                                            '<i style="text-decoration: underline;">Related source links: </i><br>' +
                                                            global_cat[indicator_code]['sourcelinks'] + '<br>' if global_cat[indicator_code]['sourcelinks'] else '',
                                                            '<i style="text-decoration: underline;">Other web links: </i><br>' + global_cat[indicator_code][
                                                                'weblinks'] + '<br>' if global_cat[indicator_code]['weblinks'] else ''
                                                            ),
                                                           datasetId=newdataset.pk)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))

                                        newvariable = Variable(name=global_cat[indicator_code]['name'], unit=global_cat[indicator_code]['unitofmeasure'] if global_cat[indicator_code]['unitofmeasure'] else '', description=global_cat[indicator_code]['description'],
                                                               code=indicator_code, timespan='1960-' + str(last_available_year), fk_dst_id=newdataset, fk_var_type_id=VariableType.objects.get(pk=4), sourceId=newsource)
                                        newvariable.save()
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        global_cat[indicator_code]['saved'] = True
                                    else:
                                        newvariable = global_cat[indicator_code]['variable_object']
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'], country_name_entity_ref[country_code].pk, newvariable.pk))
                                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []

                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        newimport = ImportHistory(import_type='wdi', import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Initial import of WDI',
                                  import_state=json.dumps({'file_hash': file_checksum(wdi_downloads_save_location + 'wdi.zip')}))
        newimport.save()

        logger.info("Import complete.")

    else:
        last_import = import_history.last()
        deleted_indicators = {}  # This is used to keep track which variables' data values were already deleted before writing new values

        if json.loads(last_import.import_state)['file_hash'] == file_checksum(wdi_downloads_save_location + 'wdi.zip'):
            logger.info('No updates available.')
            sys.exit('No updates available.')

        logger.info('New data is available.')
        available_variables = Variable.objects.filter(fk_dst_id__in=Dataset.objects.filter(fk_dst_cat_id=DatasetCategory.objects.get(name=wdi_category_name_in_db))).values('code')
        available_variables_list = []

        for each in available_variables:
            available_variables_list.append(each['code'])

        wb = load_workbook(excel_filename, read_only=True)

        series_ws = wb['Series']
        data_ws = wb['Data']
        country_ws = wb['Country']

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0  # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        global_cat = {}  # global catalog of indicators

        # data in the worksheets is not loaded into memory at once, that causes RAM to quickly fill up
        # instead, we go through each row and cell one-by-one, looking at each piece of data separately
        # this has the disadvantage of needing to traverse the worksheet several times, if we need to look up some rows/cells again

        for row in series_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        global_cat[cell.value] = {}
                        indicatordict = global_cat[cell.value]
                    if column_number == 2:
                        indicatordict['category'] = cell.value.split(':')[0]
                    if column_number == 3:
                        indicatordict['name'] = cell.value
                    if column_number == 5:
                        indicatordict['description'] = cell.value
                    if column_number == 6:
                        if cell.value:
                            indicatordict['unitofmeasure'] = cell.value
                        else:
                            if '(' not in indicatordict['name']:
                                indicatordict['unitofmeasure'] = ''
                            else:
                                indicatordict['unitofmeasure'] = indicatordict['name'][indicatordict['name'].index('('):-1].replace('(', '').replace(')','')
                    if column_number == 11:
                        if cell.value:
                            indicatordict['limitations'] = cell.value
                        else:
                            indicatordict['limitations'] = ''
                    if column_number == 12:
                        if cell.value:
                            indicatordict['sourcenotes'] = cell.value
                        else:
                            indicatordict['sourcenotes'] = ''
                    if column_number == 13:
                        if cell.value:
                            indicatordict['comments'] = cell.value
                        else:
                            indicatordict['comments'] = ''
                    if column_number == 14:
                        indicatordict['source'] = cell.value
                    if column_number == 15:
                        if cell.value:
                            indicatordict['concept'] = cell.value
                        else:
                            indicatordict['concept'] = ''
                    if column_number == 17:
                        if cell.value:
                            indicatordict['sourcelinks'] = cell.value
                        else:
                            indicatordict['sourcelinks'] = ''
                    if column_number == 18:
                        if cell.value:
                            indicatordict['weblinks'] = cell.value
                        else:
                            indicatordict['weblinks'] = ''
                    indicatordict['saved'] = False

            column_number = 0

        new_variables = []

        for key, value in global_cat.items():
            new_variables.append(key)

        vars_to_add = list(set(new_variables).difference(available_variables_list))
        newly_added_vars = list(set(new_variables).difference(available_variables_list))
        vars_to_delete = list(set(available_variables_list).difference(new_variables))

        if vars_to_delete:
            with connection.cursor() as c:
                c.execute('DELETE FROM %s WHERE fk_var_id IN (SELECT id from %s WHERE code IN (%s));' %
                      (DataValue._meta.db_table, Variable._meta.db_table, ', '.join('"{0}"'.format(w) for w in vars_to_delete)))
            logger.info('Deleting data values for variables that are not present in the new dataset: ' + ', '.join('"{0}"'.format(w) for w in vars_to_delete))

        sources_to_delete = Variable.objects.filter(code__in=vars_to_delete).values('sourceId')
        sources_to_delete_list_int = []
        for each in sources_to_delete:
            sources_to_delete_list_int.append(each['sourceId'])

        if vars_to_delete:
            Variable.objects.filter(code__in=vars_to_delete).delete()
            logger.info('Deleting variables that are not present in the new dataset: ' + ', '.join('"{0}"'.format(w) for w in vars_to_delete))

        if vars_to_delete:
            source_names_to_delete = Source.objects.filter(pk__in=sources_to_delete_list_int).values('name')
            source_names_to_delete_str = ', '.join('"{0}"'.format(w['name']) for w in source_names_to_delete)
            Source.objects.filter(pk__in=sources_to_delete_list_int).delete()
            logger.info('Deleting sources for variables that are not present in the new dataset: %s' % source_names_to_delete_str.encode('utf8'))

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if wdi_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=wdi_category_name_in_db)
            the_category.save()
            logger.info("Inserting a category %s." % wdi_category_name_in_db.encode('utf8'))

        else:
            the_category = DatasetCategory.objects.get(name=wdi_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(fk_dst_cat_id=the_category).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        wdi_categories_list = []

        for key, value in category_vars.items():
            wdi_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = DatasetSubcategory(name=key, fk_dst_cat_id=the_category)
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        cats_to_add = list(set(wdi_categories_list).difference(list(existing_subcategories_list)))
        cats_to_delete = list(set(list(existing_subcategories_list)).difference(wdi_categories_list))

        if cats_to_delete:
            Dataset.objects.filter(fk_dst_subcat_id__in=DatasetSubcategory.objects.filter(name__in=cats_to_delete, fk_dst_cat_id=the_category)).delete()
            logger.info("Deleting the datasets under the categories %s." % ', '.join('"{0}"'.format(w) for w in cats_to_delete))

        if cats_to_delete:
            DatasetSubcategory.objects.filter(name__in=cats_to_delete, fk_dst_cat_id=the_category).delete()
            logger.info("Deleting the categories %s." % ', '.join('"{0}"'.format(w) for w in cats_to_delete))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'] for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        AdditionalCountryInfo.objects.all().delete()  # We will load new additional country data now

        row_number = 0
        for row in country_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        country_code = cell.value
                    if column_number == 3:
                        country_name = cell.value
                    if column_number == 7:
                        country_special_notes = cell.value
                    if column_number == 8:
                        country_region = cell.value
                    if column_number == 9:
                        country_income_group = cell.value
                    if column_number == 24:
                        country_latest_census = cell.value
                    if column_number == 25:
                        country_latest_survey = cell.value
                    if column_number == 26:
                        country_recent_income_source = cell.value
                    if column_number == 31:
                        entity_info = AdditionalCountryInfo()
                        entity_info.country_code = country_code
                        entity_info.country_name = country_name
                        entity_info.country_wb_region = country_region
                        entity_info.country_wb_income_group = country_income_group
                        entity_info.country_special_notes = country_special_notes
                        entity_info.country_latest_census = country_latest_census
                        entity_info.country_latest_survey = country_latest_survey
                        entity_info.country_recent_income_source = country_recent_income_source
                        entity_info.save()
                        if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                            newentity = Entity.objects.get(name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                        elif country_name in existing_entities_list:
                            newentity = Entity.objects.get(name=country_name)
                        else:
                            newentity = Entity(name=country_name, validated=False)
                            newentity.save()
                            logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                        country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, fk_ent_id, fk_var_id) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []

        total_values_tracker = 0

        for category in wdi_categories_list:
            if category in cats_to_add:
                newdataset = Dataset(name='World Development Indicators - ' + category,
                                     description='This is a dataset imported by the automated fetcher',
                                     namespace='wdi', fk_dst_cat_id=the_category,
                                     fk_dst_subcat_id=DatasetSubcategory.objects.get(name=category,
                                                                                     fk_dst_cat_id=the_category))
                newdataset.save()
                logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            else:
                newdataset = Dataset.objects.get(name='World Development Indicators - ' + category, fk_dst_cat_id=DatasetCategory.objects.get(
                                                                                         name=wdi_category_name_in_db))
            row_number = 0
            for row in data_ws.rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 1:
                            country_name = cell.value
                        if column_number == 2:
                            country_code = cell.value
                        if column_number == 3:
                            indicator_name = cell.value
                        if column_number == 4:
                            indicator_code = cell.value
                        if column_number > 4 and column_number < last_available_year - 1960 + 5:
                            if cell.value:
                                data_values.append({'value': cell.value, 'year': 1960 - 5 + column_number})
                        if column_number > 4 and column_number == last_available_year - 1960 + 5:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    total_values_tracker += len(data_values)
                                    if indicator_code in vars_to_add:
                                        newsource = Source(name='World Bank – WDI: ' + global_cat[indicator_code]['name'],
                                                           description=source_template %
                                                                       (global_cat[indicator_code]['category'],
                                                                        global_cat[indicator_code]['unitofmeasure'],
                                                                        indicator_code,
                                                                        'World Bank – World Development Indicators',
                                                                        global_cat[indicator_code]['source'],
                                                                        '<i style="text-decoration: underline;">Limitations and exceptions: </i><br>' +
                                                                        global_cat[indicator_code][
                                                                            'limitations'] + '<br>' if
                                                                        global_cat[indicator_code]['limitations'] else '',
                                                                        '<i style="text-decoration: underline;">Notes from original source: </i><br>' +
                                                                        global_cat[indicator_code][
                                                                            'sourcenotes'] + '<br>' if
                                                                        global_cat[indicator_code]['sourcenotes'] else '',
                                                                        '<i style="text-decoration: underline;">General comments: </i><br>' +
                                                                        global_cat[indicator_code][
                                                                            'comments'] + '<br>' if
                                                                        global_cat[indicator_code]['comments'] else '',
                                                                        '<i style="text-decoration: underline;">Statistical concept and methodology: </i><br>' +
                                                                        global_cat[indicator_code]['concept'] + '<br>' if
                                                                        global_cat[indicator_code]['concept'] else '',
                                                                        '<i style="text-decoration: underline;">Related source links: </i><br>' +
                                                                        global_cat[indicator_code][
                                                                            'sourcelinks'] + '<br>' if
                                                                        global_cat[indicator_code]['sourcelinks'] else '',
                                                                        '<i style="text-decoration: underline;">Other web links: </i><br>' +
                                                                        global_cat[indicator_code][
                                                                            'weblinks'] + '<br>' if
                                                                        global_cat[indicator_code]['weblinks'] else ''
                                                                        ),
                                                           datasetId=newdataset.pk)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))
                                        global_cat[indicator_code]['source_object'] = newsource
                                        newvariable = Variable(name=global_cat[indicator_code]['name'],
                                                               unit=global_cat[indicator_code]['unitofmeasure'] if
                                                               global_cat[indicator_code]['unitofmeasure'] else '',
                                                               description=global_cat[indicator_code]['description'],
                                                               code=indicator_code,
                                                               timespan='1960-' + str(last_available_year),
                                                               fk_dst_id=newdataset,
                                                               fk_var_type_id=VariableType.objects.get(pk=4),
                                                               sourceId=newsource)
                                        newvariable.save()
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        vars_to_add.remove(indicator_code)
                                        global_cat[indicator_code]['saved'] = True
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                    else:
                                        if not global_cat[indicator_code]['saved']:
                                            newsource = Source.objects.get(name='World Bank – WDI: ' + Variable.objects.get(code=indicator_code).name)
                                            newsource.name = 'World Bank – WDI: ' + global_cat[indicator_code]['name']
                                            newsource.description=source_template % (global_cat[indicator_code]['category'],
                                                            global_cat[indicator_code]['unitofmeasure'],
                                                            indicator_code,
                                                            'World Bank – World Development Indicators',
                                                            global_cat[indicator_code]['source'],
                                                            '<i style="text-decoration: underline;">Limitations and exceptions: </i><br>' +
                                                            global_cat[indicator_code]['limitations'] + '<br>' if global_cat[indicator_code]['limitations'] else '',
                                                            '<i style="text-decoration: underline;">Notes from original source: </i><br>' +
                                                            global_cat[indicator_code]['sourcenotes'] + '<br>' if global_cat[indicator_code]['sourcenotes'] else '',
                                                            '<i style="text-decoration: underline;">General comments: </i><br>' + global_cat[indicator_code][
                                                                'comments'] + '<br>' if global_cat[indicator_code]['comments'] else '',
                                                            '<i style="text-decoration: underline;">Statistical concept and methodology: </i><br>' +
                                                            global_cat[indicator_code]['concept'] + '<br>' if global_cat[indicator_code]['concept'] else '',
                                                            '<i style="text-decoration: underline;">Related source links: </i><br>' +
                                                            global_cat[indicator_code]['sourcelinks'] + '<br>' if global_cat[indicator_code]['sourcelinks'] else '',
                                                            '<i style="text-decoration: underline;">Other web links: </i><br>' + global_cat[indicator_code][
                                                                'weblinks'] + '<br>' if global_cat[indicator_code]['weblinks'] else ''
                                                            )
                                            newsource.datasetId=newdataset.pk
                                            newsource.save()
                                            logger.info("Updating the source %s." % newsource.name.encode('utf8'))
                                            newvariable = Variable.objects.get(code=indicator_code)
                                            newvariable.name = global_cat[indicator_code]['name']
                                            newvariable.unit=global_cat[indicator_code]['unitofmeasure'] if global_cat[indicator_code]['unitofmeasure'] else ''
                                            newvariable.description=global_cat[indicator_code]['description']
                                            newvariable.timespan='1960-' + str(last_available_year)
                                            newvariable.fk_dst_id=newdataset
                                            newvariable.sourceId=newsource
                                            newvariable.save()
                                            global_cat[indicator_code]['variable_object'] = newvariable
                                            logger.info("Updating the variable %s." % newvariable.name.encode('utf8'))
                                            global_cat[indicator_code]['saved'] = True
                                        else:
                                            newvariable = global_cat[indicator_code]['variable_object']
                                        if indicator_code not in newly_added_vars:
                                            if not deleted_indicators.get(indicator_code, 0):
                                                with connection.cursor() as c:
                                                    c.execute(
                                                        'DELETE FROM %s WHERE fk_var_id = %s;' %
                                                        (DataValue._meta.db_table, newvariable.pk))
                                                deleted_indicators[indicator_code] = True
                                                logger.info("Deleting data values for the variable %s." % indicator_code.encode('utf8'))
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'],
                                                                       country_name_entity_ref[country_code].pk,
                                                                       newvariable.pk))
                                    if len(
                                        data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []
                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        newimport = ImportHistory(import_type='wdi', import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Imported a total of %s data values.' % total_values_tracker,
                                  import_state=json.dumps(
                                      {'file_hash': file_checksum(wdi_downloads_save_location + 'wdi.zip')}))
        newimport.save()

print("--- %s seconds ---" % (time.time() - start_time))
