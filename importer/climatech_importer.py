import sys
import os
import hashlib
import json
import logging
import unidecode
import shutil
import time
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import grapher_admin.wsgi
from openpyxl import load_workbook
from grapher_admin.models import Entity, DatasetSubcategory, DatasetCategory, Dataset, Source, Variable, VariableType, DataValue, ChartDimension
from importer.models import ImportHistory
from country_name_tool.models import CountryName
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from grapher_admin.views import write_dataset_csv


# we will use the file checksum to check if the downloaded file has changed since we last saw it
def file_checksum(filename, blocksize=2**20):
    m = hashlib.md5()
    with open(filename, "rb") as f:
        while True:
            buffer = f.read(blocksize)
            if not buffer:
                break
            m.update(buffer)
    return m.hexdigest()


def short_unit_extract(unit: str):
    common_short_units = ['$', '£', '€', '%']  # used for extracting short forms of units of measurement
    short_unit = None
    if unit:
        if ' per ' in unit:
            short_form = unit.split(' per ')[0]
            if any(w in short_form for w in common_short_units):
                for x in common_short_units:
                    if x in short_form:
                        short_unit = x
                        break
            else:
                short_unit = short_form
        elif any(x in unit for x in common_short_units):
            for y in common_short_units:
                if y in unit:
                    short_unit = y
                    break
        elif 'percentage' in unit:
            short_unit = '%'
        elif 'percent' in unit.lower():
            short_unit = '%'
        elif len(unit) < 9:  # this length is sort of arbitrary at this point, taken from the unit 'hectares'
            short_unit = unit
    return short_unit


source_description = {
    'dataPublishedBy': "World Bank Climate Change Data",
    'link': "https://data.worldbank.org/data-catalog/climate-change",
    'retrievedDate': timezone.now().strftime("%d-%B-%y")
}

climatech_zip_file_url = 'http://databank.worldbank.org/data/download/catalog/climate_change_download_0.xlsx'
climatech_downloads_save_location = settings.BASE_DIR + '/data/climatech_downloads/'

# create a directory for holding the downloads
# if the directory exists, delete it and recreate it

if not os.path.exists(climatech_downloads_save_location):
    os.makedirs(climatech_downloads_save_location)
#else:
#    shutil.rmtree(climatech_downloads_save_location)
#    os.makedirs(climatech_downloads_save_location)

logger = logging.getLogger('importer')
start_time = time.time()

climatech_category_name_in_db = 'World Bank Climate Change Data'  # set the name of the root category of all data that will be imported by this script

import_history = ImportHistory.objects.filter(import_type='climatech')

excel_filename = climatech_downloads_save_location + "climate_change_download_0.xlsx"

with transaction.atomic():
    # if climatech imports were never performed
    if not import_history:
        logger.info("This is the very first climatech data import.")

        wb = load_workbook(excel_filename, read_only=True)

        series_ws = wb['Series']
        data_ws = wb['Data']
        country_ws = wb['Country']

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0   # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        global_cat = {}  # global catalog of indicators

        # data in the worksheets is not loaded into memory at once, that causes RAM to quickly fill up
        # instead, we go through each row and cell one-by-one, looking at each piece of data separately
        # this has the disadvantage of needing to traverse the worksheet several times, if we need to look up some rows/cells again

        for row in series_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        global_cat[cell.value.upper().strip()] = {}
                        indicatordict = global_cat[cell.value.upper().strip()]
                    if column_number == 2:
                        indicatordict['name'] = cell.value
                    if column_number == 6:
                        indicatordict['category'] = cell.value
                    if column_number == 7:
                        indicatordict['description'] = cell.value
                    if column_number == 8:
                        indicatordict['source'] = cell.value
                    if column_number == 3:
                        if '(' not in indicatordict['name']:
                            indicatordict['unitofmeasure'] = ''
                        else:
                            indicatordict['unitofmeasure'] = indicatordict['name'][
                                                             indicatordict['name'].rfind('(') + 1:indicatordict[
                                                                 'name'].rfind(')')]
                    indicatordict['saved'] = False

            column_number = 0

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if climatech_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=climatech_category_name_in_db, fetcher_autocreated=True)
            the_category.save()
            logger.info("Inserting a category %s." % climatech_category_name_in_db.encode('utf8'))

        else:
            the_category = DatasetCategory.objects.get(name=climatech_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(categoryId=the_category.pk).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        climatech_categories_list = []

        for key, value in category_vars.items():
            climatech_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = DatasetSubcategory(name=key, categoryId=the_category)
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'].lower() for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        row_number = 0
        for row in country_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if cell.value:
                        if column_number == 1:
                            country_code = cell.value
                        if column_number == 2:
                            country_name = cell.value
                            if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                                newentity = Entity.objects.get(name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                            elif country_name.lower() in existing_entities_list:
                                newentity = Entity.objects.get(name=country_name)
                            else:
                                newentity = Entity(name=country_name, validated=False)
                                newentity.save()
                                logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                            country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, entityId, variableId) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []
        datasets_list = []
        for category in climatech_categories_list:
            newdataset = Dataset(name='World Bank Climate Change Data - ' + category,
                                 description='This is a dataset imported by the automated fetcher',
                                 namespace='climatech', categoryId=the_category,
                                 subcategoryId=DatasetSubcategory.objects.get(name=category, categoryId=the_category))
            newdataset.save()
            datasets_list.append(newdataset)
            logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            row_number = 0
            columns_to_years = {}
            for row in data_ws.rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        column_number += 1
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                                columns_to_years[column_number] = last_available_year
                                last_available_column = column_number
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 2:
                            country_name = cell.value
                        if column_number == 1:
                            country_code = cell.value
                        if column_number == 4:
                            indicator_name = cell.value
                        if column_number == 3:
                            indicator_code = cell.value.upper().strip()
                        if column_number > 6 and column_number <= last_available_column:
                            if cell.value or cell.value == 0:
                                try:
                                    data_values.append({'value': str(float(cell.value)), 'year': columns_to_years[column_number]})
                                except ValueError:
                                    pass
                        if column_number > 4 and column_number == last_available_column:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    if not global_cat[indicator_code]['saved']:
                                        source_description['additionalInfo'] = None
                                        source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                        if 'iea.org' in json.dumps(source_description).lower() or 'iea stat' in json.dumps(source_description).lower() or 'iea 2014' in json.dumps(source_description).lower():
                                            source_description['dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                        else:
                                            source_description['dataPublishedBy'] = "World Bank Climate Change Data"
                                        newsource = Source(name='World Bank Climate Change Data: ' + global_cat[indicator_code]['name'],
                                                           description=json.dumps(source_description),
                                                           datasetId=newdataset.pk)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))
                                        if global_cat[indicator_code]['unitofmeasure']:
                                            if len(global_cat[indicator_code]['unitofmeasure']) < 40:
                                                unit_of_measure = global_cat[indicator_code]['unitofmeasure']
                                            else:
                                                unit_of_measure = ''
                                        else:
                                            unit_of_measure = ''
                                        s_unit = short_unit_extract(unit_of_measure)
                                        newvariable = Variable(name=global_cat[indicator_code]['name'], unit=unit_of_measure, short_unit=s_unit, description=global_cat[indicator_code]['description'],
                                                               code=indicator_code, timespan='', datasetId=newdataset, variableTypeId=VariableType.objects.get(pk=4), sourceId=newsource)
                                        newvariable.save()
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        global_cat[indicator_code]['saved'] = True
                                    else:
                                        newvariable = global_cat[indicator_code]['variable_object']
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'], country_name_entity_ref[country_code].pk, newvariable.pk))
                                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []

                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        newimport = ImportHistory(import_type='climatech', import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Initial import of climatech datasets',
                                  import_state=json.dumps({'file_hash': file_checksum(excel_filename)}))
        newimport.save()
        for dataset in datasets_list:
            write_dataset_csv(dataset.pk, dataset.name, None, 'climatech_fetcher', '')
        logger.info("Import complete.")

    else:
        last_import = import_history.last()
        deleted_indicators = {}  # This is used to keep track which variables' data values were already deleted before writing new values

        if json.loads(last_import.import_state)['file_hash'] == file_checksum(excel_filename):
            logger.info('No updates available.')
            sys.exit('No updates available.')

        logger.info('New data is available.')
        available_variables = Variable.objects.filter(datasetId__in=Dataset.objects.filter(namespace='climatech'))
        available_variables_list = []

        for each in available_variables.values('code'):
            available_variables_list.append(each['code'])

        chart_dimension_vars = ChartDimension.objects.all().values('variableId').distinct()
        chart_dimension_vars_list = {item['variableId'] for item in chart_dimension_vars}
        existing_variables_ids = [item['id'] for item in available_variables.values('id')]
        existing_variables_id_code = {item['id']: item['code'] for item in available_variables.values('id', 'code')}
        existing_variables_code_id = {item['code']: item['id'] for item in available_variables.values('id', 'code')}

        vars_being_used = []  # we will not be deleting any variables that are currently being used by charts
        for each_var in existing_variables_ids:
            if each_var in chart_dimension_vars_list:
                vars_being_used.append(existing_variables_id_code[each_var])

        wb = load_workbook(excel_filename, read_only=True)

        series_ws = wb['Series']
        data_ws = wb['Data']
        country_ws = wb['Country']

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0  # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        global_cat = {}  # global catalog of indicators

        # data in the worksheets is not loaded into memory at once, that causes RAM to quickly fill up
        # instead, we go through each row and cell one-by-one, looking at each piece of data separately
        # this has the disadvantage of needing to traverse the worksheet several times, if we need to look up some rows/cells again

        for row in series_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        global_cat[cell.value.upper().strip()] = {}
                        indicatordict = global_cat[cell.value.upper().strip()]
                    if column_number == 2:
                        indicatordict['name'] = cell.value
                    if column_number == 6:
                        indicatordict['category'] = cell.value
                    if column_number == 7:
                        indicatordict['description'] = cell.value
                    if column_number == 8:
                        indicatordict['source'] = cell.value
                    if column_number == 3:
                        if '(' not in indicatordict['name']:
                            indicatordict['unitofmeasure'] = ''
                        else:
                            indicatordict['unitofmeasure'] = indicatordict['name'][
                                                             indicatordict['name'].rfind('(') + 1:indicatordict[
                                                                 'name'].rfind(')')]
                    indicatordict['saved'] = False

            column_number = 0

        new_variables = []

        for key, value in global_cat.items():
            new_variables.append(key)

        vars_to_add = list(set(new_variables).difference(available_variables_list))
        newly_added_vars = list(set(new_variables).difference(available_variables_list))
        vars_to_delete = list(set(available_variables_list).difference(new_variables))

        for each in vars_to_delete:
            if each not in vars_being_used:
                logger.info("Deleting data values for the variable: %s" % each.encode('utf8'))
                while DataValue.objects.filter(variableId__pk=existing_variables_code_id[each]).first():
                    with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                        c.execute('DELETE FROM %s WHERE variableId = %s LIMIT 10000;' %
                                  (DataValue._meta.db_table, existing_variables_code_id[each]))
                source_object = Variable.objects.get(code=each, datasetId__in=Dataset.objects.filter(namespace='climatech')).sourceId
                Variable.objects.get(code=each, datasetId__in=Dataset.objects.filter(namespace='climatech')).delete()
                logger.info("Deleting the variable: %s" % each.encode('utf8'))
                logger.info("Deleting the source: %s" % source_object.name.encode('utf8'))
                source_object.delete()

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = DatasetCategory.objects.values('name')
        existing_categories_list = {item['name'] for item in existing_categories}

        if climatech_category_name_in_db not in existing_categories_list:
            the_category = DatasetCategory(name=climatech_category_name_in_db, fetcher_autocreated=True)
            the_category.save()
            logger.info("Inserting a category %s." % climatech_category_name_in_db.encode('utf8'))

        else:
            the_category = DatasetCategory.objects.get(name=climatech_category_name_in_db)

        existing_subcategories = DatasetSubcategory.objects.filter(categoryId=the_category).values('name')
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        climatech_categories_list = []

        for key, value in category_vars.items():
            climatech_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = DatasetSubcategory(name=key, categoryId=the_category)
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        cats_to_add = list(set(climatech_categories_list).difference(list(existing_subcategories_list)))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'].lower() for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        row_number = 0
        for row in country_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if cell.value:
                        if column_number == 1:
                            country_code = cell.value
                        if column_number == 2:
                            country_name = cell.value
                            if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                                newentity = Entity.objects.get(
                                    name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                            elif country_name.lower() in existing_entities_list:
                                newentity = Entity.objects.get(name=country_name)
                            else:
                                newentity = Entity(name=country_name, validated=False)
                                newentity.save()
                                logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                            country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, entityId, variableId) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []

        total_values_tracker = 0
        dataset_id_oldname_list = []

        for category in climatech_categories_list:
            if category in cats_to_add:
                newdataset = Dataset(name='World Bank Climate Change Data - ' + category,
                                     description='This is a dataset imported by the automated fetcher',
                                     namespace='climatech', categoryId=the_category,
                                     subcategoryId=DatasetSubcategory.objects.get(name=category,
                                                                                     categoryId=the_category))
                newdataset.save()
                dataset_id_oldname_list.append({'id': newdataset.pk, 'newname': newdataset.name, 'oldname': None})
                logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            else:
                newdataset = Dataset.objects.get(name='World Bank Climate Change Data - ' + category, categoryId=DatasetCategory.objects.get(
                                                                                         name=climatech_category_name_in_db))
                dataset_id_oldname_list.append({'id': newdataset.pk, 'newname': newdataset.name, 'oldname': newdataset.name})
            row_number = 0
            columns_to_years = {}
            for row in data_ws.rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        column_number += 1
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                                columns_to_years[column_number] = last_available_year
                                last_available_column = column_number
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 2:
                            country_name = cell.value
                        if column_number == 1:
                            country_code = cell.value
                        if column_number == 4:
                            indicator_name = cell.value
                        if column_number == 3:
                            indicator_code = cell.value.upper().strip()
                        if column_number > 6 and column_number <= last_available_column:
                            if cell.value or cell.value == 0:
                                try:
                                    data_values.append(
                                        {'value': str(float(cell.value)), 'year': columns_to_years[column_number]})
                                except ValueError:
                                    pass
                        if column_number > 4 and column_number == last_available_column:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    total_values_tracker += len(data_values)
                                    if indicator_code in vars_to_add:
                                        source_description['additionalInfo'] = None
                                        source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                        if 'iea.org' in json.dumps(source_description).lower() or 'iea stat' in json.dumps(source_description).lower() or 'iea 2014' in json.dumps(source_description).lower():
                                            source_description['dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                        else:
                                            source_description['dataPublishedBy'] = "World Bank Climate Change Data"
                                        newsource = Source(name='World Bank Climate Change Data: ' + global_cat[indicator_code]['name'],
                                                           description=json.dumps(source_description),
                                                           datasetId=newdataset.pk)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))
                                        global_cat[indicator_code]['source_object'] = newsource
                                        if global_cat[indicator_code]['unitofmeasure']:
                                            if len(global_cat[indicator_code]['unitofmeasure']) < 40:
                                                unit_of_measure = global_cat[indicator_code]['unitofmeasure']
                                            else:
                                                unit_of_measure = ''
                                        else:
                                            unit_of_measure = ''
                                        s_unit = short_unit_extract(unit_of_measure)
                                        newvariable = Variable(name=global_cat[indicator_code]['name'],
                                                               unit=unit_of_measure, short_unit=s_unit,
                                                               description=global_cat[indicator_code]['description'],
                                                               code=indicator_code,
                                                               timespan='',
                                                               datasetId=newdataset,
                                                               variableTypeId=VariableType.objects.get(pk=4),
                                                               sourceId=newsource)
                                        newvariable.save()
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        vars_to_add.remove(indicator_code)
                                        global_cat[indicator_code]['saved'] = True
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                    else:
                                        if not global_cat[indicator_code]['saved']:
                                            newsource = Source.objects.get(name='World Bank Climate Change Data: ' + Variable.objects.get(code=indicator_code, datasetId__in=Dataset.objects.filter(namespace='climatech')).name)
                                            newsource.name = 'World Bank Climate Change Data: ' + global_cat[indicator_code]['name']
                                            source_description['additionalInfo'] = None
                                            source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                            if 'iea.org' in json.dumps(
                                                source_description).lower() or 'iea stat' in json.dumps(
                                                source_description).lower() or 'iea 2014' in json.dumps(
                                                source_description).lower():
                                                source_description[
                                                    'dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                            else:
                                                source_description['dataPublishedBy'] = "World Bank Climate Change Data"
                                            newsource.description=json.dumps(source_description)
                                            newsource.datasetId=newdataset.pk
                                            newsource.save()
                                            logger.info("Updating the source %s." % newsource.name.encode('utf8'))
                                            if global_cat[indicator_code]['unitofmeasure']:
                                                if len(global_cat[indicator_code]['unitofmeasure']) < 40:
                                                    unit_of_measure = global_cat[indicator_code]['unitofmeasure']
                                                else:
                                                    unit_of_measure = ''
                                            else:
                                                unit_of_measure = ''
                                            s_unit = short_unit_extract(unit_of_measure)
                                            newvariable = Variable.objects.get(code=indicator_code, datasetId__in=Dataset.objects.filter(namespace='climatech'))
                                            newvariable.name = global_cat[indicator_code]['name']
                                            newvariable.unit=unit_of_measure
                                            newvariable.short_unit = s_unit
                                            newvariable.description=global_cat[indicator_code]['description']
                                            newvariable.timespan=''
                                            newvariable.datasetId=newdataset
                                            newvariable.sourceId=newsource
                                            newvariable.save()
                                            global_cat[indicator_code]['variable_object'] = newvariable
                                            logger.info("Updating the variable %s." % newvariable.name.encode('utf8'))
                                            global_cat[indicator_code]['saved'] = True
                                        else:
                                            newvariable = global_cat[indicator_code]['variable_object']
                                        if indicator_code not in newly_added_vars:
                                            if not deleted_indicators.get(indicator_code, 0):
                                                while DataValue.objects.filter(variableId__pk=newvariable.pk).first():
                                                    with connection.cursor() as c:
                                                        c.execute(
                                                                  'DELETE FROM %s WHERE variableId = %s LIMIT 10000;' %
                                                                  (DataValue._meta.db_table, newvariable.pk))
                                                deleted_indicators[indicator_code] = True
                                                logger.info("Deleting data values for the variable %s." % indicator_code.encode('utf8'))
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'],
                                                                       country_name_entity_ref[country_code].pk,
                                                                       newvariable.pk))
                                    if len(
                                        data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []
                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        # now deleting subcategories and datasets that are empty (that don't contain any variables), if any

        all_climatech_datasets = Dataset.objects.filter(namespace='climatech')
        all_climatech_datasets_with_vars = Variable.objects.filter(datasetId__in=all_climatech_datasets).values(
            'datasetId').distinct()
        all_climatech_datasets_with_vars_dict = {item['datasetId'] for item in all_climatech_datasets_with_vars}

        for each in all_climatech_datasets:
            if each.pk not in all_climatech_datasets_with_vars_dict:
                cat_to_delete = each.subcategoryId
                logger.info("Deleting empty dataset %s." % each.name)
                logger.info("Deleting empty category %s." % cat_to_delete.name)
                each.delete()
                cat_to_delete.delete()

        newimport = ImportHistory(import_type='climatech', import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Imported a total of %s data values.' % total_values_tracker,
                                  import_state=json.dumps(
                                      {'file_hash': file_checksum(excel_filename)}))
        newimport.save()

        # now exporting csvs to the repo
        for dataset in dataset_id_oldname_list:
            write_dataset_csv(dataset['id'], dataset['newname'], dataset['oldname'], 'climatech_fetcher', '')

print("--- %s seconds ---" % (time.time() - start_time))
