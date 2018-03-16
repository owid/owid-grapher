from django.shortcuts import render
from django.http import HttpRequest, HttpResponse
from grapher_admin.models import Variable
from typing import Dict
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from importer.models import AdditionalCountryInfo


def listunwppdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace__contains='unwpp')
    datasets: Dict = {}

    for each in variables:
        dataset_full_name = each.datasetId.namespace + '<br><br>' + each.datasetId.name
        if datasets.get(dataset_full_name):
            datasets[dataset_full_name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[dataset_full_name] = []
            datasets[dataset_full_name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.unwpp.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listqogdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='qog')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.qog.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listfaodatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='faostat')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.faostat.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listclioinfradatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='clioinfra')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.clioinfra.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listgbdcausedatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='gbd_cause')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.gbdcause.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listgbdriskdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='gbd_risk')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.gbdrisk.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listunsdgdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='un_sdg')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.unsdg.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listgbdprevalencedatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='gbd_prevalence')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.gbdprevalence.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listgbdprevalencebygenderdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='gbd_prevalence_by_gender')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.gbdprevalencebygender.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listgbdmentalhealthdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='gbd_mental_health')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.gbdmentalhealth.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listilostatdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='ilostat')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.ilostat.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listunepdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='un_ep')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.unep.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listwhowashdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='who_wash')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.whowash.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listoecdstatdatasets(request: HttpRequest):
    variables = Variable.objects.filter(datasetId__namespace='oecd_stat')
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.oecdstat.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets})


def listwbdatasets(request: HttpRequest, dataset: str):
    if dataset == 'wdidatasets':
        dataset_name = 'wdi'
        dataset_title = 'World Bank WDI Datasets'
    elif dataset == 'edstatsdatasets':
        dataset_name = 'edstats'
        dataset_title = 'EdStats Datasets'
    elif dataset == 'genderstatsdatasets':
        dataset_name = 'genderstats'
        dataset_title = 'World Bank Gender Statistics Datasets'
    elif dataset == 'hnpstatsdatasets':
        dataset_name = 'hnpstats'
        dataset_title = 'World Bank Health Nutrition and Population Statistics Datasets'
    elif dataset == 'findexdatasets':
        dataset_name = 'findex'
        dataset_title = 'World Bank Global Findex Datasets'
    elif dataset == 'bbscdatasets':
        dataset_name = 'bbsc'
        dataset_title = 'World Bank Data on Statistical Capacity'
    elif dataset == 'povstatsdatasets':
        dataset_name = 'povstats'
        dataset_title = 'World Bank Poverty and Equity database'
    elif dataset == 'climatechdatasets':
        dataset_name = 'climatech'
        dataset_title = 'World Bank Climate Change Data'
    elif dataset == 'hnpqstatsdatasets':
        dataset_name = 'hnpqstats'
        dataset_title = 'World Bank Health Nutrition and Population Statistics by Wealth Quintile'
    elif dataset == 'se4alldatasets':
        dataset_name = 'se4all'
        dataset_title = 'World Bank SE4ALL database'
    elif dataset == 'aspiredatasets':
        dataset_name = 'aspire'
        dataset_title = 'World Bank The Atlas of Social Protection: Indicators of Resilience and Equity'

    variables = Variable.objects.filter(datasetId__namespace=dataset_name)
    datasets: Dict = {}

    for each in variables:
        if datasets.get(each.datasetId.subcategoryId.name):
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})
        else:
            datasets[each.datasetId.subcategoryId.name] = []
            datasets[each.datasetId.subcategoryId.name].append({'id': each.pk, 'name': each.name, 'code': each.code})

    return render(request, 'admin.wb.data.html', context={'current_user': request.user.name,
                                                           'datasets': datasets,
                                                           'dataset_title': dataset_title})


def serve_wb_country_info_xls(request: HttpRequest):
    if 'WDI_Country_info.xls' in request.path:
        filename = 'WDI_Country_info.xls'
        dataset_name = 'wdi'
    if 'EDSTATS_Country_info.xls' in request.path:
        filename = 'EDSTATS_Country_info.xls'
        dataset_name = 'edstats'
    if 'GENDERSTATS_Country_info.xls' in request.path:
        filename = 'GENDERSTATS_Country_info.xls'
        dataset_name = 'genderstats'
    if 'HNPSTATS_Country_info.xls' in request.path:
        filename = 'HNPSTATS_Country_info.xls'
        dataset_name = 'hnpstats'
    if 'FINDEX_Country_info.xls' in request.path:
        filename = 'FINDEX_Country_info.xls'
        dataset_name = 'findex'
    if 'BBSC_Country_info.xls' in request.path:
        filename = 'BBSC_Country_info.xls'
        dataset_name = 'bbsc'
    if 'POVSTATS_Country_info.xls' in request.path:
        filename = 'POVSTATS_Country_info.xls'
        dataset_name = 'povstats'
    if 'HNPQSTATS_Country_info.xls' in request.path:
        filename = 'HNPQSTATS_Country_info.xls'
        dataset_name = 'hnpqstats'
    if 'ASPIRE_Country_info.xls' in request.path:
        filename = 'ASPIRE_Country_info.xls'
        dataset_name = 'aspire'

    wb = Workbook()

    ws = wb.worksheets[0]

    all_wdi_additional_country_info = AdditionalCountryInfo.objects.filter(dataset=dataset_name)

    ws.cell(column=1, row=1, value="Country")
    ws.cell(column=2, row=1, value="Country's World Bank Region")
    ws.cell(column=3, row=1, value="Country's World Bank income group")
    ws.cell(column=4, row=1, value="Special notes")
    ws.cell(column=5, row=1, value="Latest population census")
    ws.cell(column=6, row=1, value="Latest household survey")
    ws.cell(column=7, row=1, value="Source of most recent Income and expenditure data")

    row = 2
    for each in all_wdi_additional_country_info:
        ws.cell(column=1, row=row, value="{0}".format(each.country_name))
        ws.cell(column=2, row=row, value="{0}".format(each.country_wb_region))
        ws.cell(column=3, row=row, value="{0}".format(each.country_wb_income_group))
        ws.cell(column=4, row=row, value="{0}".format(each.country_special_notes))
        ws.cell(column=5, row=row, value="{0}".format(each.country_latest_census))
        ws.cell(column=6, row=row, value="{0}".format(each.country_latest_survey))
        ws.cell(column=7, row=row, value="{0}".format(each.country_recent_income_source))

        row += 1

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return response
